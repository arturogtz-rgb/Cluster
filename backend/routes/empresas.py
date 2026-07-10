from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone
import uuid
import io
from slugify import slugify

from database import db
from auth import get_current_user, require_admin
from models import Empresa, EmpresaCreate, EmpresaUpdate, CATEGORIES

router = APIRouter()


@router.get("/empresas/plantilla")
async def download_empresa_template():
    """Generate and return an .xlsx template for importing empresas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # Main sheet with columns
    ws = wb.active
    ws.title = "Empresas"

    headers = [
        "nombre", "categoria", "descripcion", "telefono", "whatsapp",
        "direccion", "email", "facebook", "instagram", "twitter",
        "youtube", "linkedin", "website", "actividades", "latitud",
        "longitud", "destacada", "activa",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A4D2E", end_color="1A4D2E", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Example row
    example = [
        "Ecomuk Aventura Natural", "Operadora de aventura",
        "Descripcion de la empresa...", "+52 333 460 1257",
        "523334601257", "Zapopan, Jalisco", "contacto@ecomuk.com.mx",
        "https://facebook.com/ecomuk", "https://instagram.com/ecomuk", "",
        "", "", "https://ecomuk.com.mx", "Senderismo, Rappel, Kayak",
        "20.7214", "-103.4189", "TRUE", "TRUE",
    ]
    for col_idx, val in enumerate(example, 1):
        ws.cell(row=2, column=col_idx, value=val)

    # Auto-width
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(header) + 4, 15)

    # Categories reference sheet
    ws_cat = wb.create_sheet("Categorias_Validas")
    ws_cat.cell(row=1, column=1, value="Categorias validas").font = Font(bold=True)
    for i, cat in enumerate(CATEGORIES, 2):
        ws_cat.cell(row=i, column=1, value=cat)

    # Activities reference sheet
    all_actividades = await db.actividades.find({"activa": True}, {"_id": 0, "nombre": 1}).to_list(200)
    ws_act = wb.create_sheet("Actividades_Disponibles")
    ws_act.cell(row=1, column=1, value="Actividades disponibles").font = Font(bold=True)
    for i, act in enumerate(all_actividades, 2):
        ws_act.cell(row=i, column=1, value=act["nombre"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_empresas.xlsx"},
    )


@router.post("/empresas/importar")
async def import_empresas(file: UploadFile = File(...), user=Depends(require_admin)):
    """Import empresas from an .xlsx file. Validates row by row."""
    from openpyxl import load_workbook

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="El archivo no contiene datos (solo encabezados o vacio)")

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    required_fields = {"nombre", "categoria", "descripcion"}
    missing = required_fields - set(headers)
    if missing:
        raise HTTPException(status_code=400, detail=f"Columnas obligatorias faltantes: {', '.join(missing)}")

    col_map = {h: i for i, h in enumerate(headers) if h}

    importadas = 0
    actualizadas = 0
    errores = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            def get_val(col_name, default=""):
                idx = col_map.get(col_name)
                if idx is None or idx >= len(row):
                    return default
                val = row[idx]
                return str(val).strip() if val is not None else default

            nombre = get_val("nombre")
            categoria = get_val("categoria")
            descripcion = get_val("descripcion")

            if not nombre:
                errores.append({"fila": row_idx, "motivo": "Campo 'nombre' vacio"})
                continue
            if not categoria:
                errores.append({"fila": row_idx, "motivo": "Campo 'categoria' vacio"})
                continue
            if not descripcion:
                errores.append({"fila": row_idx, "motivo": "Campo 'descripcion' vacio"})
                continue

            if categoria not in CATEGORIES:
                errores.append({"fila": row_idx, "motivo": f"Categoria '{categoria}' no valida. Opciones: {', '.join(CATEGORIES)}"})
                continue

            lat_str = get_val("latitud")
            lng_str = get_val("longitud")
            latitud = None
            longitud = None
            if lat_str:
                try:
                    latitud = float(lat_str)
                except ValueError:
                    errores.append({"fila": row_idx, "motivo": f"Latitud '{lat_str}' no es un numero valido"})
                    continue
            if lng_str:
                try:
                    longitud = float(lng_str)
                except ValueError:
                    errores.append({"fila": row_idx, "motivo": f"Longitud '{lng_str}' no es un numero valido"})
                    continue

            dest_str = get_val("destacada").upper()
            destacada = dest_str in ("TRUE", "SI", "SÍ", "1", "VERDADERO")
            activa_str = get_val("activa", "TRUE").upper()
            activa = activa_str in ("TRUE", "SI", "SÍ", "1", "VERDADERO", "")

            actividades_str = get_val("actividades")
            actividades = [a.strip() for a in actividades_str.split(",") if a.strip()] if actividades_str else []

            slug = slugify(nombre, lowercase=True)
            existing = await db.empresas.find_one({"slug": slug})

            doc = {
                "nombre": nombre,
                "categoria": categoria,
                "descripcion": descripcion,
                "telefono": get_val("telefono") or None,
                "whatsapp": get_val("whatsapp") or None,
                "direccion": get_val("direccion") or None,
                "email": get_val("email") or None,
                "social_links": {
                    "facebook": get_val("facebook") or None,
                    "instagram": get_val("instagram") or None,
                    "twitter": get_val("twitter") or None,
                    "youtube": get_val("youtube") or None,
                    "linkedin": get_val("linkedin") or None,
                    "website": get_val("website") or None,
                },
                "actividades": actividades,
                "latitud": latitud,
                "longitud": longitud,
                "destacada": destacada,
                "activa": activa,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }

            if existing:
                await db.empresas.update_one({"slug": slug}, {"$set": doc})
                actualizadas += 1
            else:
                doc["id"] = str(uuid.uuid4())
                doc["slug"] = slug
                doc["logo_url"] = None
                doc["hero_url"] = None
                doc["galeria"] = []
                doc["ubicaciones_actividades"] = []
                doc["created_at"] = datetime.now(timezone.utc).isoformat()
                await db.empresas.insert_one(doc)
                importadas += 1

        except Exception as e:
            errores.append({"fila": row_idx, "motivo": str(e)})

    return {"importadas": importadas, "actualizadas": actualizadas, "errores": errores}


@router.get("/empresas/exportar")
async def export_empresas(user=Depends(require_admin)):
    """Export all empresas as .xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    empresas = await db.empresas.find({}, {"_id": 0}).sort("nombre", 1).to_list(5000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas"

    headers = [
        "nombre", "categoria", "descripcion", "telefono", "whatsapp",
        "direccion", "email", "facebook", "instagram", "twitter",
        "youtube", "linkedin", "website", "actividades", "latitud",
        "longitud", "destacada", "activa",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A4D2E", end_color="1A4D2E", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, emp in enumerate(empresas, 2):
        social = emp.get("social_links") or {}
        acts = emp.get("actividades") or []
        ws.cell(row=row_idx, column=1, value=emp.get("nombre", ""))
        ws.cell(row=row_idx, column=2, value=emp.get("categoria", ""))
        ws.cell(row=row_idx, column=3, value=emp.get("descripcion", ""))
        ws.cell(row=row_idx, column=4, value=emp.get("telefono", ""))
        ws.cell(row=row_idx, column=5, value=emp.get("whatsapp", ""))
        ws.cell(row=row_idx, column=6, value=emp.get("direccion", ""))
        ws.cell(row=row_idx, column=7, value=emp.get("email", ""))
        ws.cell(row=row_idx, column=8, value=social.get("facebook", ""))
        ws.cell(row=row_idx, column=9, value=social.get("instagram", ""))
        ws.cell(row=row_idx, column=10, value=social.get("twitter", ""))
        ws.cell(row=row_idx, column=11, value=social.get("youtube", ""))
        ws.cell(row=row_idx, column=12, value=social.get("linkedin", ""))
        ws.cell(row=row_idx, column=13, value=social.get("website", ""))
        ws.cell(row=row_idx, column=14, value=", ".join(acts))
        ws.cell(row=row_idx, column=15, value=emp.get("latitud"))
        ws.cell(row=row_idx, column=16, value=emp.get("longitud"))
        ws.cell(row=row_idx, column=17, value="TRUE" if emp.get("destacada") else "FALSE")
        ws.cell(row=row_idx, column=18, value="TRUE" if emp.get("activa") else "FALSE")

    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(header) + 4, 15)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=empresas_cluster_turismo.xlsx"},
    )


@router.get("/empresas", response_model=List[Empresa])
async def get_empresas(
    categoria: Optional[str] = None,
    busqueda: Optional[str] = None,
    destacada: Optional[bool] = None,
    activa: bool = True,
):
    query = {"activa": activa}
    if categoria:
        query["categoria"] = categoria
    if destacada is not None:
        query["destacada"] = destacada
    if busqueda:
        query["$or"] = [
            {"nombre": {"$regex": busqueda, "$options": "i"}},
            {"descripcion": {"$regex": busqueda, "$options": "i"}},
            {"actividades": {"$regex": busqueda, "$options": "i"}},
        ]
    empresas = await db.empresas.find(query, {"_id": 0}).to_list(1000)
    for empresa in empresas:
        if isinstance(empresa.get("created_at"), str):
            empresa["created_at"] = datetime.fromisoformat(empresa["created_at"])
        if isinstance(empresa.get("updated_at"), str):
            empresa["updated_at"] = datetime.fromisoformat(empresa["updated_at"])
    # Resolve activity IDs to names
    all_actividades = await db.actividades.find({}, {"_id": 0, "id": 1, "nombre": 1, "slug": 1}).to_list(200)
    act_map = {}
    for a in all_actividades:
        act_map[a["id"]] = a["nombre"]
        act_map[a.get("slug", "")] = a["nombre"]
    for empresa in empresas:
        if empresa.get("actividades"):
            empresa["actividades"] = [act_map.get(act, act) for act in empresa["actividades"]]
    return empresas


@router.get("/empresas-top-views")
async def get_top_viewed_empresas():
    empresas = await db.empresas.find(
        {"activa": True},
        {"_id": 0, "nombre": 1, "slug": 1, "views": 1, "categoria": 1, "logo_url": 1},
    ).sort("views", -1).limit(5).to_list(5)
    for e in empresas:
        e.setdefault("views", 0)
    return empresas


@router.get("/empresas-destacadas")
async def get_empresas_destacadas():
    # Priority: destacada=true first, then top by views, limit 6
    destacadas = await db.empresas.find(
        {"activa": True, "destacada": True}, {"_id": 0}
    ).sort("views", -1).to_list(20)
    for e in destacadas:
        if isinstance(e.get("created_at"), str):
            e["created_at"] = datetime.fromisoformat(e["created_at"])
        if isinstance(e.get("updated_at"), str):
            e["updated_at"] = datetime.fromisoformat(e["updated_at"])
    slugs_used = {e["slug"] for e in destacadas}
    remaining = 6 - len(destacadas)
    if remaining > 0:
        top_views = await db.empresas.find(
            {"activa": True, "slug": {"$nin": list(slugs_used)}}, {"_id": 0}
        ).sort("views", -1).limit(remaining).to_list(remaining)
        for e in top_views:
            if isinstance(e.get("created_at"), str):
                e["created_at"] = datetime.fromisoformat(e["created_at"])
            if isinstance(e.get("updated_at"), str):
                e["updated_at"] = datetime.fromisoformat(e["updated_at"])
        destacadas.extend(top_views)
    result = destacadas[:6]
    # Resolve activity IDs to names
    all_actividades = await db.actividades.find({}, {"_id": 0, "id": 1, "nombre": 1, "slug": 1}).to_list(200)
    act_map = {}
    for a in all_actividades:
        act_map[a["id"]] = a["nombre"]
        act_map[a.get("slug", "")] = a["nombre"]
    for e in result:
        if e.get("actividades"):
            e["actividades"] = [act_map.get(act, act) for act in e["actividades"]]
    return result


@router.get("/empresas/{slug}", response_model=Empresa)
async def get_empresa(slug: str):
    empresa = await db.empresas.find_one({"slug": slug}, {"_id": 0})
    if not empresa:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    if isinstance(empresa.get("created_at"), str):
        empresa["created_at"] = datetime.fromisoformat(empresa["created_at"])
    if isinstance(empresa.get("updated_at"), str):
        empresa["updated_at"] = datetime.fromisoformat(empresa["updated_at"])
    # Resolve activity IDs to names
    if empresa.get("actividades"):
        all_actividades = await db.actividades.find({}, {"_id": 0, "id": 1, "nombre": 1, "slug": 1}).to_list(200)
        act_map = {}
        for a in all_actividades:
            act_map[a["id"]] = a["nombre"]
            act_map[a.get("slug", "")] = a["nombre"]
        empresa["actividades"] = [act_map.get(act, act) for act in empresa["actividades"]]
    await db.empresas.update_one({"slug": slug}, {"$inc": {"views": 1}})
    return empresa


@router.post("/empresas", response_model=Empresa)
async def create_empresa(data: EmpresaCreate, user=Depends(get_current_user)):
    empresa_dict = data.model_dump()
    empresa_dict["slug"] = slugify(data.nombre, lowercase=True)
    existing = await db.empresas.find_one({"slug": empresa_dict["slug"]})
    if existing:
        empresa_dict["slug"] = f"{empresa_dict['slug']}-{str(uuid.uuid4())[:8]}"
    empresa = Empresa(**empresa_dict)
    doc = empresa.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    doc["updated_at"] = doc["updated_at"].isoformat()
    await db.empresas.insert_one(doc)
    return empresa


@router.put("/empresas/{slug}", response_model=Empresa)
async def update_empresa(slug: str, data: EmpresaUpdate, user=Depends(get_current_user)):
    from fastapi import HTTPException
    empresa = await db.empresas.find_one({"slug": slug}, {"_id": 0})
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if "nombre" in update_data:
        new_slug = slugify(update_data["nombre"], lowercase=True)
        if new_slug != slug:
            existing = await db.empresas.find_one({"slug": new_slug})
            if existing:
                new_slug = f"{new_slug}-{str(uuid.uuid4())[:8]}"
            update_data["slug"] = new_slug
    update_data["updated_at"] = datetime.now().isoformat()
    if "social_links" in update_data and update_data["social_links"]:
        update_data["social_links"] = (
            update_data["social_links"].model_dump()
            if hasattr(update_data["social_links"], "model_dump")
            else update_data["social_links"]
        )
    await db.empresas.update_one({"slug": slug}, {"$set": update_data})
    updated = await db.empresas.find_one({"slug": update_data.get("slug", slug)}, {"_id": 0})
    if isinstance(updated.get("created_at"), str):
        updated["created_at"] = datetime.fromisoformat(updated["created_at"])
    if isinstance(updated.get("updated_at"), str):
        updated["updated_at"] = datetime.fromisoformat(updated["updated_at"])
    return updated


@router.delete("/empresas/{slug}")
async def delete_empresa(slug: str, user=Depends(get_current_user)):
    from fastapi import HTTPException
    result = await db.empresas.delete_one({"slug": slug})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    return {"message": "Empresa eliminada"}
